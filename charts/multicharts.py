#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

# sbk_charts :  Storage Benchmark Kit - Charts
from collections import OrderedDict
from openpyxl.chart import Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ordered_set import OrderedSet

import charts.constants as constants
from charts.charts import SbkCharts


class SbkMultiCharts(SbkCharts):
    def __init__(self, version, file):
        super().__init__(version, file)

    def check_time_units(self):
        ret = OrderedSet()
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ret.add(self.get_time_unit(self.wb[name]))
        if len(ret) > 1:
            print("ERROR: Multiple Time unit are preset in " + self.file + " " + str(ret))
            return False
        print("Time Unit : " + ''.join(ret))
        return True

    def get_actions_storage_map(self):
        ret = OrderedDict()
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                action = self.get_action_name(self.wb[name])
                if action not in ret:
                    ret[action] = OrderedSet()
                ret[action].add(self.get_storage_name(self.wb[name]))
        return ret

    def create_summary_sheet(self):
        BLACK = 'FF000000'
        WHITE = 'FFFFFFFF'
        RED = 'FFFF0000'
        DARKRED = 'FF800000'
        BLUE = 'FF0000FF'
        DARKBLUE = 'FF000080'
        GREEN = 'FF00FF00'
        DARKGREEN = 'FF008000'
        YELLOW = 'FFFFFF00'
        DARKYELLOW = 'FF808000'

        acts = self.get_actions_storage_map()
        sheet = self.wb.create_sheet("Summary")
        row = 7
        col = 7
        sheet.column_dimensions[get_column_letter(col)].width = 25
        sheet.column_dimensions[get_column_letter(col + 1)].width = 50
        cell = sheet.cell(row, col + 1)
        cell.value = "SBK Charts "
        cell.font = Font(size="47", bold=True, color=DARKBLUE)
        cell.alignment = Alignment(horizontal='center')
        row += 1
        cell = sheet.cell(row, col + 1)
        cell.value = "SBK Version : " + self.version
        cell.font = Font(size="27", bold=True, color=DARKYELLOW)
        cell.alignment = Alignment(horizontal='center')
        row += 1
        drivers = OrderedSet()
        for values in acts.values():
            drivers.update(values)
        text = "Performance Analysis of Storage Drivers :  " + ", ".join(drivers)
        cell = sheet.cell(row, col)
        cell.value = text
        cell.font = Font(size="27", bold=True, color=RED)
        row += 1
        cell = sheet.cell(row, col)
        cell.value = "Time Unit"
        cell.font = Font(size="18", bold=False, color=BLUE)
        cell = sheet.cell(row, col + 1)
        cell.value = self.get_time_unit(self.wb[constants.R_PREFIX + "1"])
        cell.font = Font(size="18", bold=False, color=BLACK)
        row += 1
        for i, key in enumerate(acts):
            cell = sheet.cell(row + i, col)
            cell.value = key
            text = key
            cell.font = Font(size="18", bold=False, color=DARKGREEN)
            cell = sheet.cell(row + i, col + 1)
            cell.value = ", ".join(acts[key])
            cell.font = Font(size="18", bold=False, color=DARKRED)
            text += " : " + cell.value
            print(text)

    def create_all_latency_compare_graphs(self):
        charts, sheets = [], []
        for i in range(self.n_latency_charts):
            charts.append(self.create_latency_line_graph("Latency Variations"))
            sheets.append(self.wb.create_sheet("Latencies-" + str(i + 1)))
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                latency_series = self.get_latency_series(ws, prefix)
                for x in latency_series:
                    for i, g in enumerate(self.latency_groups):
                        if x in g:
                            charts[i].append(latency_series[x])
        for i, ch in enumerate(charts):
            ch.width = 70
            ch.height = 70
            sheets[i].add_chart(ch)

    def create_multi_latency_compare_graphs(self):
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                super().create_latency_compare_graphs(ws, prefix)

    def create_multi_latency_graphs(self):
        charts = OrderedDict()
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                latency_series = self.get_latency_series(ws, prefix)
                for x in latency_series:
                    if x not in charts:
                        charts[x] = self.create_latency_line_graph(x + " Variations")
                    charts[x].append(latency_series[x])
        for x in charts:
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(charts[x])

    def create_total_multi_latency_percentile_graphs(self):
        title = "Total Percentiles"
        for i, names_list in enumerate(self.slc_percentile_names):
            chart = self.create_line_chart(title, "Percentiles", "Latency time in " + self.time_unit, 25, 50)
            x_labels = False
            for name in self.wb.sheetnames:
                if self.is_tnum_sheet(name):
                    ws = self.wb[name]
                    prefix = name + "_" + self.get_storage_name(ws)
                    latency_series = self.get_latency_percentile_series(ws, prefix, names_list)
                    for x in latency_series:
                        chart.append(latency_series[x])
                    if x_labels is False:
                        latency_cols = self.get_latency_percentile_columns(ws)
                        percentile_names = Reference(ws, min_col=latency_cols[names_list[0]], min_row=1,
                                                     max_col=latency_cols[names_list[-1]], max_row=1)
                        chart.set_categories(percentile_names)
                        x_labels = True
            sheet = self.wb.create_sheet("Total_Percentiles_" + str(i + 1))
            sheet.add_chart(chart)

    def create_multi_throughput_mb_graph(self, ):
        chart = self.create_line_chart("Throughput Variations in Mega Bytes / Seconds",
                                       "Intervals", "Throughput in MB/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_MB")
        sheet.add_chart(chart)

    def create_multi_throughput_records_graph(self):
        chart = self.create_line_chart("Throughput Variations in Records / Seconds",
                                       "Intervals", "Throughput in Records/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_records_series(ws, prefix))
                chart.append(self.get_throughput_read_request_records_series(ws, prefix))
                chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_Records")
        sheet.add_chart(chart)

    def create_multi_write_read_records_graph(self):
        chart = self.create_line_chart("Write and Read Records Variations",
                                       "Intervals", "Write and Read Records", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_records_series(ws, prefix))
                chart.append(self.get_read_request_records_series(ws, prefix))
                chart.append(self.get_write_response_pending_records_series(ws, prefix))
                chart.append(self.get_read_response_pending_records_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_records_series(ws, prefix))
                chart.append(self.get_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Write_Read_Records")
        sheet.add_chart(chart)

    def create_multi_write_read_mb_graph(self):
        chart = self.create_line_chart("Write and Read MBs Variations",
                                       "Intervals", "Write and Read MBs", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_rnum_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_mb_series(ws, prefix))
                chart.append(self.get_read_request_mb_series(ws, prefix))
                chart.append(self.get_write_response_pending_mb_series(ws, prefix))
                chart.append(self.get_read_response_pending_mb_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_mb_series(ws, prefix))
                chart.append(self.get_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Write_Read_MB")
        sheet.add_chart(chart)

    def create_total_mb_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Mega Bytes " + action, action, "MB", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_mb_series(ws, prefix))
                chart.append(self.get_read_request_mb_series(ws, prefix))
                chart.append(self.get_write_response_pending_mb_series(ws, prefix))
                chart.append(self.get_read_response_pending_mb_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_mb_series(ws, prefix))
                chart.append(self.get_mb_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_MB")
            sheet.add_chart(chart)

    def create_total_throughput_mb_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Throughput Variations in Mega Bytes / Seconds",
                                                  action, "Total Throughput in MB/Sec", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_mb_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Throughput_MB")
            sheet.add_chart(chart)

    def create_total_throughput_records_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Throughput Variations in Records / Seconds",
                                                  action, "Total Throughput in Records/Sec", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_records_series(ws, prefix))
                chart.append(self.get_throughput_read_request_records_series(ws, prefix))
                chart.append(self.get_throughput_records_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Throughput_Records")
            sheet.add_chart(chart)

    def create_total_avg_latency_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Average Latency Comparison",
                                                  action, "Total Average Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_avg_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Avg_Latency")
            sheet.add_chart(chart)

    def create_total_min_latency_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Min Latency Comparison",
                                                  action, "Total Min Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_min_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Min_Latency")
            sheet.add_chart(chart)

    def create_total_max_latency_compare_graph(self):
        chart = None
        for name in self.wb.sheetnames:
            if self.is_tnum_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Max Latency Comparison",
                                                  action, "Total Max Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_max_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Max_Latency")
            sheet.add_chart(chart)

    def create_graphs(self):
        if self.check_time_units():
            self.create_summary_sheet()
            self.create_multi_throughput_mb_graph()
            self.create_multi_throughput_records_graph()
            self.create_all_latency_compare_graphs()
            self.create_multi_latency_compare_graphs()
            self.create_multi_latency_graphs()
            self.create_multi_write_read_records_graph()
            self.create_multi_write_read_mb_graph()
            self.create_total_multi_latency_percentile_graphs()
            self.create_total_mb_compare_graph()
            self.create_total_throughput_mb_compare_graph()
            self.create_total_throughput_records_compare_graph()
            self.create_total_min_latency_compare_graph()
            self.create_total_avg_latency_compare_graph()
            self.create_total_max_latency_compare_graph()
            self.wb.save(self.file)
            print("file : %s updated with graphs" % self.file)
