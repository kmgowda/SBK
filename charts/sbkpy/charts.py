#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

# sbk_charts :  Storage Benchmark Kit - Charts

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
import sbkpy.constants as constants


class SbkCharts:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.time_unit = self.get_time_unit(self.wb[constants.R_PREFIX + "1"])
        self.n_latency_charts = 4
        self.latency_groups = [
            ["Percentile_10", "Percentile_20", "Percentile_25", "Percentile_30", "Percentile_40", "Percentile_50"],
            ["Percentile_50", "AvgLatency"],
            ["Percentile_50", "Percentile_60", "Percentile_70", "Percentile_75", "Percentile_80", "Percentile_90"],
            ["Percentile_92.5", "Percentile_95", "Percentile_97.5", "Percentile_99",
             "Percentile_99.25", "Percentile_99.5", "Percentile_99.75", "Percentile_99.9",
             "Percentile_99.95", "Percentile_99.99"]]

    def get_columns_from_worksheet(self, ws):
        ret = OrderedDict()
        for cell in ws[1]:
            if cell.value:
                ret[cell.value] = cell.column
        return ret

    def get_latency_columns(self, ws):
        columns = self.get_columns_from_worksheet(ws)
        ret = OrderedDict()
        ret['AvgLatency'] = columns['AvgLatency']
        ret['MaxLatency'] = columns['MaxLatency']
        for key in columns.keys():
            if key.startswith("Percentile_"):
                ret[key] = columns[key]
        return ret

    def get_time_unit(self, ws):
        names = self.get_columns_from_worksheet(ws)
        return ws.cell(row=2, column=names['LatencyTimeUnit']).value

    def get_storage_name(self, ws):
        names = self.get_columns_from_worksheet(ws)
        return ws.cell(row=2, column=names['Storage']).value

    def create_line_chart(self, title, x_title, y_title, height, width):
        chart = LineChart()
        # set the title of the chart
        chart.title = title
        # set the title of the x-axis
        chart.x_axis.title = x_title
        # set the title of the y-axis
        chart.y_axis.title = y_title
        chart.height = height
        chart.width = width
        return chart

    def create_latency_line_graph(self, title):
        return self.create_line_chart(title, "Intervals", "Latency time in " + self.time_unit, 25, 50)

    def get_latency_series(self, ws, ws_name):
        latencies = self.get_latency_columns(ws)
        data_series = OrderedDict()
        for x in latencies:
            data_series[x] = Series(Reference(ws, min_col=latencies[x], min_row=2,
                                              max_col=latencies[x], max_row=ws.max_row),
                                    title=ws_name + "-" + x)
        return data_series

    def get_throughput_mb_series(self, ws, ws_name):
        cols = self.get_columns_from_worksheet(ws)
        return Series(Reference(ws, min_col=cols["MB/Sec"], min_row=2,
                                max_col=cols["MB/Sec"], max_row=ws.max_row),
                      title=ws_name + "-MB/Sec")

    def get_throughput_records_series(self, ws, ws_name):
        cols = self.get_columns_from_worksheet(ws)
        return Series(Reference(ws, min_col=cols["Records/Sec"], min_row=2,
                                max_col=cols["Records/Sec"], max_row=ws.max_row),
                      title=ws_name + "-Records/Sec")

    def create_latency_compare_graphs(self, ws, prefix):
        charts, sheets = [], []
        for i in range(self.n_latency_charts):
            charts.append(self.create_latency_line_graph("Latency Variations"))
            sheets.append(self.wb.create_sheet("Latencies-" + prefix + "-" + str(i + 1)))
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            for i, g in enumerate(self.latency_groups):
                if x in g:
                    charts[i].append(latency_series[x])
        for i, ch in enumerate(charts):
            sheets[i].add_chart(ch)

    def create_latency_graphs(self, ws, prefix):
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            chart = self.create_latency_line_graph(x + " Variations")
            # adding data
            chart.append(latency_series[x])
            # add chart to the sheet
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(chart)

    def create_throughput_mb_graph(self, ws, prefix):
        chart = self.create_line_chart("Throughput Variations in Mega Bytes / Seconds",
                                       "Intervals", "Throughput in MB/Sec", 25, 50)
        # adding data
        chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("MB_Sec")
        sheet.add_chart(chart)

    def create_throughput_records_graph(self, ws, prefix):
        chart = self.create_line_chart("Throughput Variations in Records / Seconds",
                                       "Intervals", "Throughput in Records/Sec", 25, 50)
        # adding data
        chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Records_Sec")
        sheet.add_chart(chart)

    def create_graphs(self):
        r_name = constants.R_PREFIX + "1"
        t_name = constants.T_PREFIX + "1"
        ws1 = self.wb[r_name]
        prefix = r_name + self.get_storage_name(ws1)
        self.create_throughput_mb_graph(ws1, prefix)
        self.create_throughput_records_graph(ws1, prefix)
        self.create_latency_compare_graphs(ws1, prefix)
        self.create_latency_graphs(ws1, prefix)
        self.wb.save(self.file)


class SbkMultiCharts(SbkCharts):
    def __init__(self, file):
        super().__init__(file)

    def create_multi_latency_compare_graphs(self):
        for name in self.wb.sheetnames:
            if name.startswith(constants.R_PREFIX):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                super().create_latency_compare_graphs(ws, prefix)

    def create_multi_latency_graphs(self):
        charts = OrderedDict()
        for name in self.wb.sheetnames:
            if name.startswith(constants.R_PREFIX):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                latency_series = self.get_latency_series(ws, prefix)
                for x in latency_series:
                    if x not in charts:
                        charts[x] = self.create_latency_line_graph(x + " Variations")
                    charts[x].append(latency_series[x])
        for x in charts:
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(charts[x])

    def create_multi_throughput_mb_graph(self, ):
        chart = self.create_line_chart("Throughput Variations in Mega Bytes / Seconds",
                                       "Intervals", "Throughput in MB/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if name.startswith(constants.R_PREFIX):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_MB")
        sheet.add_chart(chart)

    def create_multi_throughput_records_graph(self):
        chart = self.create_line_chart("Throughput Variations in Records / Seconds",
                                       "Intervals", "Throughput in Records/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if name.startswith(constants.R_PREFIX):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_Records")
        sheet.add_chart(chart)

    def create_graphs(self):
        self.create_multi_throughput_mb_graph()
        self.create_multi_throughput_records_graph()
        self.create_multi_latency_compare_graphs()
        self.create_multi_latency_graphs()
        self.wb.save(self.file)
