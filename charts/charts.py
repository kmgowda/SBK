#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

# sbk_charts :  Storage Benchmark Kit - Charts
import re
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series

import charts.constants as constants


class SbkCharts:
    def __init__(self, version, file):
        self.version = version
        self.file = file
        self.wb = load_workbook(self.file)
        self.time_unit = self.get_time_unit(self.wb[constants.R_PREFIX + "1"])
        self.n_latency_charts = 5
        self.latency_groups = [
            ["MinLatency", "Percentile_5"],
            ["Percentile_5", "Percentile_10", "Percentile_20", "Percentile_25", "Percentile_30",
             "Percentile_40", "Percentile_50"],
            ["Percentile_50", "AvgLatency"],
            ["Percentile_50", "Percentile_60", "Percentile_70", "Percentile_75", "Percentile_80", "Percentile_90"],
            ["Percentile_92.5", "Percentile_95", "Percentile_97.5", "Percentile_99",
             "Percentile_99.25", "Percentile_99.5", "Percentile_99.75", "Percentile_99.9",
             "Percentile_99.95", "Percentile_99.99"]]
        self.slc_percentile_names = [["Percentile_5", "Percentile_10", "Percentile_20", "Percentile_25",
                                      "Percentile_30", "Percentile_40", "Percentile_50"],
                                     ["Percentile_50", "Percentile_60", "Percentile_70", "Percentile_75",
                                      "Percentile_80", "Percentile_90", "Percentile_92.5", "Percentile_95",
                                      "Percentile_97.5", "Percentile_99", "Percentile_99.25", "Percentile_99.5",
                                      "Percentile_99.75", "Percentile_99.9", "Percentile_99.95", "Percentile_99.99"]]

    def is_rnum_sheet(self, name):
        return re.match("^" + constants.R_PREFIX + "\d+$", name)

    def is_tnum_sheet(self, name):
        return re.match("^" + constants.T_PREFIX + "\d+$", name)

    def get_columns_from_worksheet(self, ws):
        ret = OrderedDict()
        for cell in ws[1]:
            if cell.value:
                ret[cell.value] = cell.column
        return ret

    def get_latency_percentile_columns(self, ws):
        columns = self.get_columns_from_worksheet(ws)
        ret = OrderedDict()
        for key in columns.keys():
            if key.startswith("Percentile_"):
                ret[key] = columns[key]
        return ret

    def get_latency_columns(self, ws):
        columns = self.get_columns_from_worksheet(ws)
        ret = OrderedDict()
        ret['AvgLatency'] = columns['AvgLatency']
        ret['MinLatency'] = columns['MinLatency']
        ret['MaxLatency'] = columns['MaxLatency']
        ret.update(self.get_latency_percentile_columns(ws))
        return ret

    def get_time_unit(self, ws):
        names = self.get_columns_from_worksheet(ws)
        return str(ws.cell(row=2, column=names['LatencyTimeUnit']).value).upper()

    def get_storage_name(self, ws):
        names = self.get_columns_from_worksheet(ws)
        return str(ws.cell(row=2, column=names['Storage']).value).upper()

    def get_action_name(self, ws):
        names = self.get_columns_from_worksheet(ws)
        return str(ws.cell(row=2, column=names['Action']).value)

    def __add_chart_attriuutes(self, chart, title, x_title, y_title, height, width):
        # set the title of the chart
        chart.title = title
        # set the title of the x-axis
        chart.x_axis.title = x_title
        # set the title of the y-axis
        chart.y_axis.title = y_title
        chart.height = height
        chart.width = width

    def create_line_chart(self, title, x_title, y_title, height, width):
        chart = LineChart()
        self.__add_chart_attriuutes(chart, title, x_title, y_title, height, width)
        return chart

    def create_bar_chart(self, title, x_title, y_title, height, width):
        chart = BarChart()
        self.__add_chart_attriuutes(chart, title, x_title, y_title, height, width)
        return chart

    def create_latency_line_graph(self, title):
        return self.create_line_chart(title, "Intervals", "Latency time in " + self.time_unit, 25, 50)

    def get_latency_series(self, ws, ws_name):
        latencies = self.get_latency_columns(ws)
        data_series = OrderedDict()
        for x in latencies:
            data_series[x] = Series(Reference(ws, min_col=latencies[x], min_row=2,
                                              max_col=latencies[x], max_row=ws.max_row),
                                    title=ws_name + "-" + x)
        return data_series

    def get_latency_percentile_series(self, ws, ws_name, names_list):
        latencies = self.get_latency_percentile_columns(ws)
        data_series = OrderedDict()
        min_col = latencies[names_list[0]]
        max_col = latencies[names_list[-1]]
        for r in range(2, ws.max_row + 1):
            data_series[r] = Series(Reference(ws, min_col=min_col, min_row=r, max_col=max_col, max_row=r),
                                    title=ws_name + "_" + str(r))
        return data_series

    def get_column_series(self, ws, ws_name, column_name):
        cols = self.get_columns_from_worksheet(ws)
        return Series(Reference(ws, min_col=cols[column_name], min_row=2,
                                max_col=cols[column_name], max_row=ws.max_row),
                      title=ws_name + "-" + column_name)

    def get_throughput_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "MB/Sec")

    def get_throughput_write_request_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteRequestMB/Sec")

    def get_throughput_read_request_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "ReadRequestMB/Sec")

    def get_throughput_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "Records/Sec")

    def get_throughput_write_request_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteRequestRecords/Sec")

    def get_throughput_read_request_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "ReadRequestRecords/Sec")

    def get_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "Records")

    def get_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "MB")

    def get_write_request_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteRequestMB")

    def get_write_response_pending_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteResponsePendingMB")

    def get_write_response_pending_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteResponsePendingRecords")

    def get_read_request_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "ReadRequestMB")

    def get_read_response_pending_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "ReadResponsePendingMB")

    def get_read_response_pending_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "ReadResponsePendingRecords")

    def get_write_read_request_pending_mb_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteReadRequestPendingMB")

    def get_write_read_request_pending_records_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "WriteReadRequestPendingRecords")

    def get_avg_latency_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "AvgLatency")

    def get_min_latency_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "MinLatency")

    def get_max_latency_series(self, ws, ws_name):
        return self.get_column_series(ws, ws_name, "MaxLatency")

    def create_latency_compare_graphs(self, ws, prefix):
        charts, sheets = [], []
        for i in range(self.n_latency_charts):
            charts.append(self.create_latency_line_graph("Latency Variations"))
            sheets.append(self.wb.create_sheet("Latencies-" + prefix + "-" + str(i + 1)))
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            for i, g in enumerate(self.latency_groups):
                if x in g:
                    charts[i].append(latency_series[x])
        for i, ch in enumerate(charts):
            sheets[i].add_chart(ch)

    def create_latency_graphs(self, ws, prefix):
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            chart = self.create_latency_line_graph(x + " Variations")
            # adding data
            chart.append(latency_series[x])
            # add chart to the sheet
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(chart)

    def create_total_latency_percentile_graphs(self, ws, prefix):
        title = "Total Percentiles"
        latency_cols = self.get_latency_percentile_columns(ws)
        for i, percentile_names in enumerate(self.slc_percentile_names):
            chart = self.create_line_chart(title, "Percentiles", "Latency time in " + self.time_unit, 25, 50)
            latency_series = self.get_latency_percentile_series(ws, prefix, percentile_names)
            for x in latency_series:
                chart.append(latency_series[x])
            # Add x-axis labels
            percentiles = Reference(ws, min_col=latency_cols[percentile_names[0]], min_row=1,
                                    max_col=latency_cols[percentile_names[-1]], max_row=1)
            chart.set_categories(percentiles)
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Percentiles_" + str(i + 1))
            sheet.add_chart(chart)

    def create_throughput_mb_graph(self, ws, prefix):
        chart = self.create_line_chart("Throughput Variations in Mega Bytes / Seconds",
                                       "Intervals", "Throughput in MB/Sec", 25, 50)
        # adding data
        chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
        chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
        chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("MB_Sec")
        sheet.add_chart(chart)

    def create_throughput_records_graph(self, ws, prefix):
        chart = self.create_line_chart("Throughput Variations in Records / Seconds",
                                       "Intervals", "Throughput in Records/Sec", 25, 50)
        # adding data
        chart.append(self.get_throughput_write_request_records_series(ws, prefix))
        chart.append(self.get_throughput_read_request_records_series(ws, prefix))
        chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Records_Sec")
        sheet.add_chart(chart)

    def create_graphs(self):
        r_name = constants.R_PREFIX + "1"
        r_ws = self.wb[r_name]
        r_prefix = r_name + self.get_storage_name(r_ws)
        t_name = constants.T_PREFIX + "1"
        t_ws = self.wb[t_name]
        t_prefix = t_name + self.get_storage_name(t_ws)
        self.create_throughput_mb_graph(r_ws, r_prefix)
        self.create_throughput_records_graph(r_ws, r_prefix)
        self.create_latency_compare_graphs(r_ws, r_prefix)
        self.create_latency_graphs(r_ws, r_prefix)
        self.create_total_latency_percentile_graphs(t_ws, t_prefix)
        self.wb.save(self.file)
