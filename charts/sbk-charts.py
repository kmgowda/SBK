#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

# SBK-Charts :  Storage Benchmark Kit - Charts

import argparse
from collections import OrderedDict

import openpyxl
import pandas
import xlsxwriter
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter


class SbkCharts:
    def __init__(self, iFile, oFile):
        self.iFile = iFile
        self.oFile = oFile

    def create_sheets(self):
        df = pandas.read_csv(self.iFile)
        header = df.columns.values
        wb = xlsxwriter.Workbook(self.oFile)
        ws1 = wb.add_worksheet("R-1")
        ws2 = wb.add_worksheet("T-1")
        for c, h in enumerate(header):
            ws1.set_column(c, c, len(h))
            ws2.set_column(c, c, len(h))
            ws1.write(0, c, h)
            ws2.write(0, c, h)

        r1 = 1
        r2 = 1
        for row in df.iterrows():
            if row[1]['Type'] == 'Total':
                for c, h in enumerate(header):
                    col_size = len(str(row[1][h])) + 1
                    if col_size > len(h):
                        ws2.set_column(c, c, col_size)
                    ws2.write(r2, c, row[1][h])
                r2 += 1
            else:
                for c, h in enumerate(header):
                    col_size = len(str(row[1][h])) + 1
                    if col_size > len(h):
                        ws1.set_column(c, c, col_size)
                    ws1.write(r1, c, row[1][h])
                r1 += 1
        wb.close()
        print("xlsx file %s created" % self.oFile)

    def get_columns_from_worksheet(self, ws):
        return {
            cell.value: {
                'letter': get_column_letter(cell.column),
                'number': cell.column
            } for cell in ws[1] if cell.value
        }

    def get_latency_columns(self, ws):
        colnames = self.get_columns_from_worksheet(ws)
        ret = OrderedDict()
        ret['AvgLatency'] = colnames['AvgLatency']
        ret['MaxLatency'] = colnames['MaxLatency']
        for key in colnames.keys():
            if key.startswith("Percentile_"):
                ret[key] = colnames[key]
        return ret

    def get_time_unit(self, ws):
        colnames = self.get_columns_from_worksheet(ws)
        return ws.cell(row=2, column=colnames['LatencyTimeUnit']['number']).value

    def create_latency_graphs(self, wb, ws, time_unit):
        latencies = self.get_latency_columns(ws)
        charts = [LineChart(), LineChart(), LineChart()]

        for ch in charts:
            # set the title of the chart
            ch.title = " Percentile Variations"

            # set the title of the x-axis
            ch.x_axis.title = " Intervals "

            # set the title of the y-axis
            ch.y_axis.title = " Latency Time in " + time_unit

            ch.height = 25
            ch.width = 50

        tmpws = [wb.create_sheet("Latencies-1"), wb.create_sheet("Latencies-2"), wb.create_sheet("Latencies-3")]

        groups = [
            ["Percentile_10", "Percentile_20", "Percentile_25", "Percentile_30", "Percentile_40", "Percentile_50"],
            ["Percentile_60", "Percentile_70", "Percentile_75", "Percentile_80", "Percentile_90"],
            ["Percentile_92.5", "Percentile_95", "Percentile_97.5", "Percentile_99",
             "Percentile_99.25", "Percentile_99.5", "Percentile_99.75", "Percentile_99.9",
             "Percentile_99.95", "Percentile_99.90"]]

        for x in latencies:
            data_series = Series(Reference(ws, min_col=latencies[x]['number'], min_row=2,
                                           max_col=latencies[x]['number'], max_row=ws.max_row), title=x)

            for i, g in enumerate(groups):
                if x in g:
                    charts[i].append(data_series)

            chart = LineChart()

            # adding data
            chart.append(data_series)

            # set the title of the chart
            chart.title = x + " Variations"

            # set the title of the x-axis
            chart.x_axis.title = " Intervals "

            # set the title of the y-axis
            chart.y_axis.title = " Latency Time in " + time_unit

            chart.height = 20
            chart.width = 40

            # add chart to the sheet
            newws = wb.create_sheet(x)
            newws.add_chart(chart)
        for i, ch in enumerate(charts):
            tmpws[i].add_chart(ch)

    def create_throughput_MB_graph(self, wb, ws):
        cols = self.get_columns_from_worksheet(ws)
        data_series = Series(Reference(ws, min_col=cols["MB/Sec"]['number'], min_row=2,
                                       max_col=cols["MB/Sec"]['number'], max_row=ws.max_row), title="MB/Sec")

        chart = LineChart()

        # adding data
        chart.append(data_series)

        # set the title of the chart
        chart.title = " Throughput Variations in Mega Bytes / Seconds"

        # set the title of the x-axis
        chart.x_axis.title = "Intervals"

        # set the title of the y-axis
        chart.y_axis.title = "Throughput in MB/Sec"

        chart.height = 20
        chart.width = 40

        # add chart to the sheet
        newws = wb.create_sheet("MB_Sec")
        newws.add_chart(chart)

    def create_throughput_records_graph(self, wb, ws):
        cols = self.get_columns_from_worksheet(ws)

        data_series = Series(Reference(ws, min_col=cols["Records/Sec"]['number'], min_row=2,
                                       max_col=cols["Records/Sec"]['number'], max_row=ws.max_row), title="Records/Sec")

        chart = LineChart()

        # adding data
        chart.append(data_series)

        # set the title of the chart
        chart.title = " Throughput Variations in Records / Second"

        # set the title of the x-axis
        chart.x_axis.title = " Intervals "

        # set the title of the y-axis
        chart.y_axis.title = "Throughput in Records/Sec"

        chart.height = 20
        chart.width = 40

        # add chart to the sheet
        newws = wb.create_sheet("Records_Sec")
        newws.add_chart(chart)

    def create_graphs(self):
        self.create_sheets()
        wb = openpyxl.load_workbook(self.oFile)
        ws1 = wb["R-1"]
        ws2 = wb["T-1"]
        """
        for row in ws1.iter_rows():
            lt = []
            for cell in row:
                lt.append(cell.value)
            print(lt)
        """
        # print(self.get_columns_from_worksheet(ws1))
        self.create_throughput_MB_graph(wb, ws1)
        self.create_throughput_records_graph(wb, ws1)
        self.create_latency_graphs(wb, ws1, self.get_time_unit(ws1))
        wb.save(self.oFile)


def main():
    parser = argparse.ArgumentParser(description='sbk charts')
    parser.add_argument('-i', '--ifile', help='Input CSV file', required=True)
    parser.add_argument('-o', '--ofile', help='Output xlsx file', default="out.xlsx")
    args = parser.parse_args()
    charts = SbkCharts(args.ifile, args.ofile)
    print('Input file is ', charts.iFile)
    print('Output file is ', charts.oFile)
    charts.create_graphs()


if __name__ == "__main__":
    main()
